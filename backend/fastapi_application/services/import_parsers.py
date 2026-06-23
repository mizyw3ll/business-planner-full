"""File format parsers for business plan import."""

from __future__ import annotations

import asyncio
import csv
import io
import re

from openpyxl import load_workbook


def parse_csv(text: str) -> dict:
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    plan_title = "Imported Plan"
    plan_description = ""
    blocks_data: list[dict] = []

    for row in rows:
        if len(row) >= 2:
            if row[0] == "Plan Title":
                plan_title = row[1]
            elif row[0] == "Description":
                plan_description = row[1]
            elif row[0] == "Block Order":
                break

    header_found = False
    for row in rows:
        if not header_found:
            if row and row[0] == "Block Order":
                header_found = True
            continue
        if len(row) >= 4 and row[0]:
            try:
                blocks_data.append(
                    {
                        "block_order": int(row[0]),
                        "title": row[1],
                        "block_type": row[2],
                        "content": row[3],
                    }
                )
            except ValueError:
                continue

    return {"title": plan_title, "description": plan_description, "blocks": blocks_data}


async def parse_xlsx(content: bytes) -> dict:
    def _parse(content: bytes) -> list:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        return list(ws.iter_rows(values_only=True))

    all_rows = await asyncio.to_thread(_parse, content)

    plan_title = "Imported Plan"
    plan_description = ""
    blocks_data: list[dict] = []

    header_labels = ("Порядок", "Название блока", "Тип блока", "Содержание")
    english_labels = ("Block Order", "Block Title", "Block Type", "Content")

    header_row_idx = None
    for idx, row in enumerate(all_rows):
        if row and (row[:4] == header_labels or row[:4] == english_labels):
            header_row_idx = idx
            break

    if header_row_idx is not None:
        if all_rows and all_rows[0]:
            plan_title = str(all_rows[0][0] or "Imported Plan")
        if len(all_rows) > 1 and all_rows[1] and all_rows[1][0]:
            plan_description = str(all_rows[1][0] or "")
        for row in all_rows[header_row_idx + 1 :]:
            if not row or row[0] is None:
                continue
            try:
                blocks_data.append(
                    {
                        "block_order": int(row[0]),
                        "title": row[1] if row[1] else "Untitled",
                        "block_type": row[2] if row[2] else "general",
                        "content": row[3] if row[3] else "",
                    }
                )
            except (ValueError, TypeError):
                continue
    else:
        for row in all_rows:
            if row and row[0] == "Plan Title":
                plan_title = row[1] if row[1] else "Imported Plan"
            elif row and row[0] == "Description":
                plan_description = row[1] if row[1] else ""
            elif row and row[0] == "Block Order":
                block_rows = all_rows[all_rows.index(row) + 1 :]
                for br in block_rows:
                    if br and br[0] is not None:
                        try:
                            blocks_data.append(
                                {
                                    "block_order": int(br[0]),
                                    "title": br[1] if br[1] else "Untitled",
                                    "block_type": br[2] if br[2] else "general",
                                    "content": br[3] if br[3] else "",
                                }
                            )
                        except (ValueError, TypeError):
                            continue
                break

    return {"title": plan_title, "description": plan_description, "blocks": blocks_data}


def parse_html(text: str) -> dict:
    plan_title = "Imported Plan"
    plan_description = ""
    blocks_data: list[dict] = []

    h1_match = re.search(r"<h1[^>]*>(.*?)</h1>", text, re.S)
    if h1_match:
        plan_title = re.sub(r"<[^>]+>", "", h1_match.group(1)).strip()

    h3_pattern = re.compile(r"<h3[^>]*>(.*?)</h3>", re.S)
    content_div_pattern = re.compile(r"<div[^>]*style=\"margin-top:\s*12px[^\"]*\">(.*?)</div>", re.S)

    h3_matches = list(h3_pattern.finditer(text))
    div_matches = list(content_div_pattern.finditer(text))

    for idx, (hm, dm) in enumerate(zip(h3_matches, div_matches)):
        title = re.sub(r"<[^>]+>", "", hm.group(1)).strip()
        content_text = re.sub(r"<[^>]+>", "", dm.group(1)).strip()
        if title:
            blocks_data.append(
                {
                    "block_order": idx + 1,
                    "title": title,
                    "block_type": "general",
                    "content": content_text,
                }
            )

    return {"title": plan_title, "description": plan_description, "blocks": blocks_data}


async def parse_pdf(content: bytes) -> dict:
    from pypdf import PdfReader

    def _parse(content: bytes) -> str:
        reader = PdfReader(io.BytesIO(content))
        text_parts = []
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        return "\n".join(text_parts)

    full_text = await asyncio.to_thread(_parse, content)

    plan_title = "Imported Plan"
    plan_description = ""
    blocks_data: list[dict] = []

    lines = [line.strip() for line in full_text.split("\n") if line.strip()]
    if lines:
        plan_title = lines[0]
    remaining = "\n".join(lines[1:]) if len(lines) > 1 else ""
    if remaining:
        blocks_data.append(
            {
                "block_order": 1,
                "title": "Импортированный контент",
                "block_type": "text",
                "content": remaining,
            }
        )

    return {"title": plan_title, "description": plan_description, "blocks": blocks_data}
