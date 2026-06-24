import asyncio
import csv
import html
import io
from functools import partial
from typing import Annotated, Literal

from core.models import BusinessPlan
from fastapi import APIRouter, Depends, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .block_serializer import EXPORT_HTML_STYLES, block_type_label, serialize_block_html, serialize_block_plain
from .dependencies import business_plan_by_id
from .pdf_export import generate_plan_pdf

export_router = APIRouter()


@export_router.get("/{plan_id}/export")
async def export_plan(
    plan: Annotated[BusinessPlan, Depends(business_plan_by_id)],
    format: Literal["html", "xlsx", "csv", "pdf"] = Query("html"),
):
    if format == "pdf":
        pdf_bytes = await asyncio.to_thread(generate_plan_pdf, plan)
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=plan_{plan.id}.pdf"},
        )

    if format == "html":
        blocks_html = ""
        for block in sorted(plan.blocks, key=lambda b: b.block_order):
            rc = block.rich_content if isinstance(block.rich_content, dict) else None
            content_html = serialize_block_html(block.block_type, block.content or "", rc)
            title_esc = html.escape(block.title, quote=True)
            type_esc = html.escape(block_type_label(block.block_type), quote=True)
            blocks_html += f"""
            <div style="margin-bottom: 24px; padding: 16px; border: 1px solid #e5e7eb; border-radius: 8px;">
                <h3 style="margin: 0 0 8px; font-size: 18px; color: #111827;">{title_esc}</h3>
                <p style="margin: 0; font-size: 14px; color: #6b7280; text-transform: capitalize;">{type_esc}</p>
                <div style="margin-top: 12px; font-size: 14px; color: #374151;">{content_html}</div>
            </div>
            """

        title_esc = html.escape(plan.title, quote=True)
        desc_esc = html.escape(plan.description or "", quote=True)
        created_esc = html.escape(plan.created_at.strftime("%Y-%m-%d"), quote=True)  # type: ignore[attr-defined]
        html_content = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>{title_esc}</title>
<style>{EXPORT_HTML_STYLES}</style>
</head>
<body>
    <h1>{title_esc}</h1>
    <p class="meta">{desc_esc}<br>Создан: {created_esc}</p>
    {blocks_html}
</body>
</html>"""

        return HTMLResponse(
            content=html_content,
            headers={"Content-Disposition": f"attachment; filename=plan_{plan.id}.html"},
        )

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Plan Title", plan.title])
        writer.writerow(["Description", plan.description or ""])
        writer.writerow(["Created", plan.created_at.strftime("%Y-%m-%d")])  # type: ignore[attr-defined]
        writer.writerow([])
        writer.writerow(["Block Order", "Block Title", "Block Type", "Content"])
        for block in sorted(plan.blocks, key=lambda b: b.block_order):
            rc = block.rich_content if isinstance(block.rich_content, dict) else None
            content = serialize_block_plain(block.block_type, block.content or "", rc)
            writer.writerow(
                [
                    block.block_order,
                    block.title,
                    block.block_type,
                    content,
                ]
            )
        output.seek(0)
        data = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=plan_{plan.id}.csv"},
        )

    def _build_xlsx() -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Business Plan"

        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        header_font = Font(bold=True, color="ffffff", size=11)
        thin_border = Border(
            left=Side(style="thin", color="d1d5db"),
            right=Side(style="thin", color="d1d5db"),
            top=Side(style="thin", color="d1d5db"),
            bottom=Side(style="thin", color="d1d5db"),
        )

        ws.merge_cells("A1:E1")
        ws["A1"] = plan.title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="left")

        if plan.description:
            ws.merge_cells("A2:E2")
            ws["A2"] = plan.description
            ws["A2"].font = Font(size=10, color="6b7280")

        ws.merge_cells("A3:E3")
        ws["A3"] = f"Создан: {plan.created_at.strftime('%Y-%m-%d')}"  # type: ignore[attr-defined]
        ws["A3"].font = Font(size=10, color="9ca3af")

        ws.append([])
        headers = ["Порядок", "Название блока", "Тип блока", "Содержание (текст)", "Содержание (HTML)"]
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for block in sorted(plan.blocks, key=lambda b: b.block_order):
            rc = block.rich_content if isinstance(block.rich_content, dict) else None
            plain = serialize_block_plain(block.block_type, block.content or "", rc)
            html_content = serialize_block_html(block.block_type, block.content or "", rc)
            row_data = [block.block_order, block.title, block.block_type, plain, html_content]
            ws.append(row_data)
            for col_idx in range(1, 6):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in range(1, 6):
            max_len = 0
            from openpyxl.utils import get_column_letter

            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                val = str(row[0] or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        xlsx_output = io.BytesIO()
        wb.save(xlsx_output)
        xlsx_output.seek(0)
        return xlsx_output

    xlsx_output = await asyncio.to_thread(_build_xlsx)
    return StreamingResponse(
        xlsx_output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=plan_{plan.id}.xlsx"},
    )
