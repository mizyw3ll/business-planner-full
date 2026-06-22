import asyncio
import csv
import io

from core.models import BusinessPlan, PlanBlock, Template, User, db_helper
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from api.api_v1.auth.fastapi_users import current_active_user

from .block_serializer import ALL_BLOCK_TYPES, build_import_rich_content
from .schemas import BusinessPlanRead

import_router = APIRouter()


@import_router.post("/import", response_model=BusinessPlanRead, status_code=status.HTTP_201_CREATED)
async def import_plan(
    file: UploadFile = File(...),
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(db_helper.session_getter),
):
    content = await file.read()
    filename = (file.filename or "").lower()

    blocks_data = []
    plan_title = "Imported Plan"
    plan_description = ""

    if filename.endswith(".csv"):
        content_str = content.decode("utf-8")
        reader = csv.reader(io.StringIO(content_str))
        rows = list(reader)

        for row in rows:
            if len(row) >= 2:
                if row[0] == "Plan Title":
                    plan_title = row[1]
                elif row[0] == "Description":
                    plan_description = row[1]
                elif row[0] == "Block Order":
                    break

        header_found = False
        for row in rows:
            if not header_found:
                if row and row[0] == "Block Order":
                    header_found = True
                continue
            if len(row) >= 4 and row[0]:
                try:
                    blocks_data.append(
                        {
                            "block_order": int(row[0]),
                            "title": row[1],
                            "block_type": row[2],
                            "content": row[3],
                        }
                    )
                except ValueError:
                    continue

    elif filename.endswith((".xlsx", ".xls")):
        def _parse_xlsx(content: bytes) -> list[dict]:
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            return all_rows  # type: ignore[return-value]

        all_rows = await asyncio.to_thread(_parse_xlsx, content)

        header_row_idx = None
        header_labels = ("Порядок", "Название блока", "Тип блока", "Содержание")
        english_labels = ("Block Order", "Block Title", "Block Type", "Content")

        for idx, row in enumerate(all_rows):
            if row and row[:4] == header_labels or row and row[:4] == english_labels:
                header_row_idx = idx
                break

        if header_row_idx is not None:
            if all_rows and all_rows[0]:
                plan_title = str(all_rows[0][0] or "Imported Plan")
            if len(all_rows) > 1 and all_rows[1] and all_rows[1][0]:
                plan_description = str(all_rows[1][0] or "")
            for row in all_rows[header_row_idx + 1 :]:
                if not row or row[0] is None:
                    continue
                try:
                    blocks_data.append(
                        {
                            "block_order": int(row[0]),
                            "title": row[1] if row[1] else "Untitled",
                            "block_type": row[2] if row[2] else "general",
                            "content": row[3] if row[3] else "",
                        }
                    )
                except (ValueError, TypeError):
                    continue
        else:
            for row in all_rows:
                if row and row[0] == "Plan Title":
                    plan_title = row[1] if row[1] else "Imported Plan"
                elif row and row[0] == "Description":
                    plan_description = row[1] if row[1] else ""
                elif row and row[0] == "Block Order":
                    block_rows = all_rows[all_rows.index(row) + 1 :]
                    for br in block_rows:
                        if br and br[0] is not None:
                            try:
                                blocks_data.append(
                                    {
                                        "block_order": int(br[0]),
                                        "title": br[1] if br[1] else "Untitled",
                                        "block_type": br[2] if br[2] else "general",
                                        "content": br[3] if br[3] else "",
                                    }
                                )
                            except (ValueError, TypeError):
                                continue
                    break
    elif filename.endswith(".html"):
        import re

        content_str = content.decode("utf-8")

        h1_match = re.search(r"<h1[^>]*>(.*?)</h1>", content_str, re.S)
        if h1_match:
            plan_title = re.sub(r"<[^>]+>", "", h1_match.group(1)).strip()

        h3_pattern = re.compile(r"<h3[^>]*>(.*?)</h3>", re.S)
        content_div_pattern = re.compile(r"<div[^>]*style=\"margin-top:\s*12px[^\"]*\">(.*?)</div>", re.S)

        h3_matches = list(h3_pattern.finditer(content_str))
        div_matches = list(content_div_pattern.finditer(content_str))

        for idx, (hm, dm) in enumerate(zip(h3_matches, div_matches)):
            title = re.sub(r"<[^>]+>", "", hm.group(1)).strip()
            content_text = re.sub(r"<[^>]+>", "", dm.group(1)).strip()
            if title:
                blocks_data.append(
                    {
                        "block_order": idx + 1,
                        "title": title,
                        "block_type": "general",
                        "content": content_text,
                    }
                )

    elif filename.endswith(".pdf"):
        from pypdf import PdfReader

        def _parse_pdf(content: bytes) -> str:
            reader = PdfReader(io.BytesIO(content))  # type: ignore[assignment]
            text_parts = []
            for page in reader.pages:  # type: ignore[attr-defined]
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            return "\n".join(text_parts)

        full_text = await asyncio.to_thread(_parse_pdf, content)

        lines = [line.strip() for line in full_text.split("\n") if line.strip()]
        if lines:
            plan_title = lines[0]
        remaining = "\n".join(lines[1:]) if len(lines) > 1 else ""
        if remaining:
            blocks_data.append(
                {
                    "block_order": 1,
                    "title": "Импортированный контент",
                    "block_type": "text",
                    "content": remaining,
                }
            )
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Неподдерживаемый формат файла. Используйте CSV, XLSX, HTML или PDF.",
        )

    plan = BusinessPlan(
        title=plan_title,
        description=plan_description,
        user_id=user.id,
    )
    session.add(plan)
    await session.flush()

    for block_data in blocks_data:
        text_content: str = block_data.get("content") or ""  # type: ignore[assignment]
        raw_type: str = block_data.get("block_type") or "general"  # type: ignore[assignment]
        block_type = raw_type if raw_type in ALL_BLOCK_TYPES else "general"
        rich_content = build_import_rich_content(block_type, text_content)

        block = PlanBlock(
            business_plan_id=plan.id,
            title=block_data["title"],
            content=block_data["content"],
            block_type=block_type,
            block_order=block_data["block_order"],
            rich_content=rich_content,
            media_attachments=[],
        )
        session.add(block)

    await session.commit()

    stmt = (
        select(BusinessPlan)
        .where(BusinessPlan.id == plan.id)
        .options(
            selectinload(BusinessPlan.blocks).selectinload(PlanBlock.tags),
            selectinload(BusinessPlan.blocks).selectinload(PlanBlock.linked_financial_charts),
            selectinload(BusinessPlan.tags),
        )
    )
    result = await session.execute(stmt)
    return result.scalar_one()


@import_router.post(
    "/from-template/{template_id}", response_model=BusinessPlanRead, status_code=status.HTTP_201_CREATED
)
async def create_plan_from_template(
    template_id: int,
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(db_helper.session_getter),
):
    template = await session.get(Template, template_id)
    if not template or not template.is_public:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Шаблон не найден")

    plan = BusinessPlan(
        title=template.title,
        description=template.description,
        user_id=user.id,
    )
    session.add(plan)
    await session.flush()

    for idx, block_data in enumerate(template.blocks):
        block = PlanBlock(
            business_plan_id=plan.id,
            title=block_data.get("title", "Untitled"),
            content=block_data.get("content", ""),
            block_type=block_data.get("block_type", "general"),
            block_order=idx,
            rich_content=block_data.get("rich_content", {}),
            media_attachments=block_data.get("media_attachments", []),
        )
        session.add(block)

    await session.commit()

    stmt = (
        select(BusinessPlan)
        .where(BusinessPlan.id == plan.id)
        .options(
            selectinload(BusinessPlan.blocks).selectinload(PlanBlock.tags),
            selectinload(BusinessPlan.blocks).selectinload(PlanBlock.linked_financial_charts),
            selectinload(BusinessPlan.tags),
        )
    )
    result = await session.execute(stmt)
    return result.scalar_one()
