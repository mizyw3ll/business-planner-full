import csv
import io
from typing import Annotated, Literal

from core.models import (
    FinancialChart,
    User,
    db_helper,
)
from fastapi import (
    APIRouter,
    Depends,
    Query,
    status,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field
from services.analytics import compute_chart_analytics
from sqlalchemy.ext.asyncio import AsyncSession

from api.api_v1.auth.fastapi_users import current_active_user
from api.api_v1.financial.chart_point.schemas import ChartPointRead

from . import crud
from .dependencies import financial_chart_by_id, financial_chart_owner_only
from .schemas import (
    FinancialChartAnalyticsRead,
    FinancialChartCreate,
    FinancialChartListItemRead,
    FinancialChartRead,
    FinancialChartUpdate,
)

router = APIRouter()


@router.get("", response_model=list[FinancialChartListItemRead])
async def get_my_charts(
    include_points: bool = Query(False),
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(db_helper.session_getter),
):
    if include_points:
        return await crud.get_user_charts(
            session=session,
            user_id=user.id,
        )
    return await crud.get_user_charts_light(
        session=session,
        user_id=user.id,
    )


class ChartPointsBatchRequest(BaseModel):
    chart_ids: list[int] = Field(default_factory=list)


class ChartPointsBatchItem(BaseModel):
    chart_id: int
    points: list[ChartPointRead]


@router.post("/points/batch", response_model=list[ChartPointsBatchItem])
async def get_chart_points_batch(
    body: ChartPointsBatchRequest,
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(db_helper.session_getter),
):
    unique_ids = list(dict.fromkeys(body.chart_ids))
    grouped = await crud.get_chart_points_batch(
        session=session,
        user_id=user.id,
        chart_ids=unique_ids,
    )
    return [ChartPointsBatchItem(chart_id=cid, points=grouped.get(cid, [])) for cid in unique_ids]


@router.post(
    "",
    response_model=FinancialChartRead,
    status_code=status.HTTP_201_CREATED,
)
async def create_chart(
    chart_in: FinancialChartCreate,
    user: User = Depends(current_active_user),
    session: AsyncSession = Depends(db_helper.session_getter),
):
    return await crud.create_chart(
        session=session,
        chart_in=chart_in,
        user_id=user.id,
    )


@router.get("/{chart_id}", response_model=FinancialChartRead)
async def get_chart(
    chart: Annotated[FinancialChart, Depends(financial_chart_by_id)],
):
    return chart


@router.get("/{chart_id}/analytics", response_model=FinancialChartAnalyticsRead)
async def get_chart_analytics(
    chart: Annotated[FinancialChart, Depends(financial_chart_by_id)],
    include_series: bool = Query(True),
):
    return FinancialChartAnalyticsRead(**compute_chart_analytics(chart, include_series=include_series))


@router.patch("/{chart_id}", response_model=FinancialChartRead)
async def update_chart(
    chart_update: FinancialChartUpdate,
    chart: Annotated[FinancialChart, Depends(financial_chart_owner_only)],
    session: AsyncSession = Depends(db_helper.session_getter),
):
    return await crud.update_chart(
        session=session,
        chart=chart,
        chart_update=chart_update,
    )


@router.delete("/{chart_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_chart(
    chart: Annotated[FinancialChart, Depends(financial_chart_owner_only)],
    session: AsyncSession = Depends(db_helper.session_getter),
):
    await crud.delete_chart(
        session=session,
        chart=chart,
    )


@router.get("/{chart_id}/export")
async def export_chart(
    chart: Annotated[FinancialChart, Depends(financial_chart_by_id)],
    format: Literal["xlsx", "csv"] = Query("xlsx"),
):
    points = sorted(chart.chart_points, key=lambda p: p.date)  # type: ignore[arg-type,return-value]
    total_income = sum(float(p.amount) for p in points if p.type == "income")
    total_expense = sum(float(p.amount) for p in points if p.type == "expense")

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["date", "type", "amount", "description"])
        for p in points:
            writer.writerow(
                [
                    p.date.isoformat(),  # type: ignore[attr-defined]
                    p.type,
                    str(p.amount),
                    p.description or "",
                ]
            )
        writer.writerow([])
        writer.writerow(["Summary", "", "", ""])
        writer.writerow(["Total Income", "", str(total_income), ""])
        writer.writerow(["Total Expense", "", str(total_expense), ""])
        writer.writerow(["Net", "", str(total_income - total_expense), ""])
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=chart_{chart.id}.csv"},
        )

    # XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart Data"

    header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=11)
    summary_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="d1d5db"),
        right=Side(style="thin", color="d1d5db"),
        top=Side(style="thin", color="d1d5db"),
        bottom=Side(style="thin", color="d1d5db"),
    )

    ws.merge_cells("A1:D1")
    ws["A1"] = chart.title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Currency ID: {chart.currency_id}"
    ws["A2"].font = Font(size=10, color="6b7280")

    ws.append([])
    headers = ["Дата", "Тип", "Сумма", "Описание"]
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for p in points:
        ws.append([p.date.isoformat(), p.type, float(p.amount), p.description or ""])  # type: ignore[attr-defined]
        for col_idx in range(1, 5):
            ws.cell(row=ws.max_row, column=col_idx).border = thin_border

    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="Итого").font = Font(bold=True, size=12)
    ws.cell(row=summary_row, column=1).fill = summary_fill
    ws.cell(row=summary_row, column=1).border = thin_border
    for col_idx in range(2, 5):
        ws.cell(row=summary_row, column=col_idx).fill = summary_fill
        ws.cell(row=summary_row, column=col_idx).border = thin_border

    for label, value in [
        ("Доходы", total_income),
        ("Расходы", total_expense),
        ("Чистая прибыль", total_income - total_expense),
    ]:
        ws.append([label, value])
        for col_idx in range(1, 5):
            ws.cell(row=ws.max_row, column=col_idx).border = thin_border

    for col_idx in range(1, 5):
        max_len = 0
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = str(row[0] or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    xlsx_output = io.BytesIO()
    wb.save(xlsx_output)
    xlsx_output.seek(0)
    return StreamingResponse(
        xlsx_output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=chart_{chart.id}.xlsx"},
    )
